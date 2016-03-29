#coding:utf-8
import sys
reload(sys)
sys.setdefaultencoding( "utf-8" )

from openpyxl import load_workbook
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
import json
import urllib2

ip = '10.1.30.53'
baseParams = "appid=10710&deviceType=3&appVersion=1&deviceId=123&sign=dasdf&"
token = 'aa'
mobile = '18825242861'
'''
wb = load_workbook('test.xlsx')
ws = wb.active

ws['D3'].value = 'pass'
ws['D3'].font = Font(color=colors.GREEN)
wb.save('test.xlsx')
'''
def getToken():
	global token
	global baseParams
	params = 'verificationCode=999999&mobile='+mobile
	url ='/v2/user/signin'
	value = baseParams + params
	respjson = json.loads(httpPost(url,value))
	token = respjson['data']['token']
	token = token.replace('+','%252B')
	print token
	baseParams = baseParams + 'token=' + token + '&'
	print baseParams

'''
发起http请求处理
'''	
def httpPost(uri,params):
	url='http://' + ip + uri
	'''
	urlencode方式发送
	#values={'appid':'1710','deviceType':'3','appVersion':'1','deviceId':'1321dasdfds','sign':'dsfds','verificationCode':'999999'}
	#values["mobile"]=mobile
	#data = urllib.urlencode(values)
	'''
	#此处用的是字符串直接提交
	values = baseParams + params
	req = urllib2.Request(url,values)
	response = urllib2.urlopen(req)
	return response.read()


def readParams(filename,rowNum):
	testCase = ws.cell(row=rowNum,column=1).value
	testUri = ws.cell(row=rowNum,column=2).value
	testParams = ws.cell(row=rowNum,column=3).value
	return testCase,testUri,testParams

def writeResult(rowNum,testResult,resp):
	if testResult=='pass':
		ws.cell(row=x,column=4).font = Font(color = colors.GREEN)
		ws.cell(row=x,column=5).value = ''
	else:
		ws.cell(row=x,column=4).font = Font(color = colors.RED)
		ws.cell(row=x,column=5).value = resp
	ws.cell(row=x,column=4).value = testResult

filename = 'test.xlsx'
getToken()

wb = load_workbook(filename)
ws = wb.active
rowNum=len(ws.rows)
print rowNum
if rowNum>2:
	for x in xrange(3,rowNum+1):
		testCase = ws.cell(row=x,column=1).value
		testUri = ws.cell(row=x,column=2).value
		testParams = ws.cell(row=x,column=3).value
		rspjson = json.loads(httpPost(testUri,testParams))
		if rspjson['code'] == 0:
			print "pass"
			writeResult(x,'pass',rspjson['message'])
		else:
			print "fail"
			writeResult(x,'fail',rspjson['message'])
			print rspjson['message']

wb.save('test.xlsx')
